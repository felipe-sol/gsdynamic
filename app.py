from flask import Flask
from flask import render_template
from flask import request
from flask import jsonify
import csv
import networkx as nx


from services.graph_service import build_graph
from io import StringIO, BytesIO
from flask import Response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


from services.dynamic_programming import optimize_mission
from services.data_service import get_objects
from services.graph_service import build_graph
from services.dijkstra_service import find_best_route
from services.montecarlo_service import run_monte_carlo
from services.alert_service import get_alerts
from services.prediction_service import predict_collisions
from services.greedy_service import greedy_priority

app = Flask(__name__)


@app.route("/")
def dashboard():

    objects = get_objects()

    satellites = len(
        [o for o in objects
         if o["type"] == "satellite"]
    )

    debris = len(
        [o for o in objects
         if o["type"] == "debris"]
    )

    return render_template(
        "dashboard.html",
        satellites=satellites,
        debris=debris
    )

@app.route("/greedy")
def greedy():
    priorities = greedy_priority()

    top_priority = priorities[0] if priorities else None

    total_score = round(
        sum(obj["priority_score"] for obj in priorities),
        2
    )

    return render_template(
        "greedy.html",
        priorities=priorities,
        top_priority=top_priority,
        total_score=total_score
    )

@app.route("/radar")
def radar():
    objects = get_objects()
    return render_template(
        "radar.html",
        objects=objects
    )

@app.route("/prediction")
def prediction():

    predictions = predict_collisions()

    return render_template(
        "prediction.html",
        predictions=predictions
    )

@app.route("/history")
def history():
    events = [
        {
            "time": "08:15:22",
            "type": "Alerta",
            "severity": "Alta",
            "title": "DEBRIS-A entrou em zona crítica",
            "description": "Objeto com risco orbital elevado próximo a satélites em LEO.",
            "action": "Executar Monte Carlo e priorizar no algoritmo guloso."
        },
        {
            "time": "09:42:10",
            "type": "Manobra",
            "severity": "Média",
            "title": "STARLINK-1001 recalculou trajetória",
            "description": "Rota orbital ajustada para reduzir proximidade com detrito.",
            "action": "Registrar alteração no relatório executivo."
        },
        {
            "time": "10:30:48",
            "type": "Dijkstra",
            "severity": "Baixa",
            "title": "Caminho mínimo calculado",
            "description": "Algoritmo de Dijkstra encontrou rota orbital de menor custo.",
            "action": "Validar rota segura no painel de grafos."
        },
        {
            "time": "11:05:33",
            "type": "Sistema",
            "severity": "Baixa",
            "title": "Relatório exportado",
            "description": "Relatório CSV/Excel gerado com dados e insights operacionais.",
            "action": "Disponibilizar arquivo para análise."
        }
    ]

    return render_template(
        "history.html",
        events=events
    )

@app.route("/control-center")
def control_center():
    objects = get_objects()

    satellites = len([o for o in objects if o["type"] == "satellite"])
    debris = len([o for o in objects if o["type"] == "debris"])
    critical = len([o for o in objects if o["risk"] >= 70])

    return render_template(
        "control_center.html",
        satellites=satellites,
        debris=debris,
        critical=critical
    )

@app.route("/map")
def orbital_map():
    objects = get_objects()

    mapped_objects = []

    for obj in objects:
        item = obj.copy()

        if obj["altitude"] <= 2000:
            item["orbit_region"] = "LEO"
        elif obj["altitude"] <= 35786:
            item["orbit_region"] = "MEO"
        else:
            item["orbit_region"] = "GEO"

        mapped_objects.append(item)

    leo = len([o for o in mapped_objects if o["orbit_region"] == "LEO"])
    meo = len([o for o in mapped_objects if o["orbit_region"] == "MEO"])
    geo = len([o for o in mapped_objects if o["orbit_region"] == "GEO"])

    return render_template(
        "map.html",
        objects=mapped_objects,
        leo=leo,
        meo=meo,
        geo=geo
    )

@app.route("/alerts")
def alerts():

    alerts = get_alerts()

    return render_template(
        "alerts.html",
        alerts=alerts
    )

@app.route("/tracking")
def tracking():
    objects = get_objects()

    search = request.args.get("search", "").lower()
    obj_type = request.args.get("type", "")
    risk_filter = request.args.get("risk", "")
    sort_by = request.args.get("sort", "risk_desc")

    filtered = []

    for obj in objects:
        matches_search = search in obj["name"].lower()

        matches_type = (
            obj_type == ""
            or obj["type"] == obj_type
        )

        if obj["risk"] >= 70:
            risk_level = "high"
        elif obj["risk"] >= 40:
            risk_level = "medium"
        else:
            risk_level = "low"

        matches_risk = (
            risk_filter == ""
            or risk_filter == risk_level
        )

        if matches_search and matches_type and matches_risk:
            filtered.append(obj)

    if sort_by == "risk_desc":
        filtered.sort(key=lambda o: o["risk"], reverse=True)
    elif sort_by == "risk_asc":
        filtered.sort(key=lambda o: o["risk"])
    elif sort_by == "altitude_desc":
        filtered.sort(key=lambda o: o["altitude"], reverse=True)
    elif sort_by == "altitude_asc":
        filtered.sort(key=lambda o: o["altitude"])
    elif sort_by == "name":
        filtered.sort(key=lambda o: o["name"])

    satellites = len([
        o for o in filtered
        if o["type"] == "satellite"
    ])

    debris = len([
        o for o in filtered
        if o["type"] == "debris"
    ])

    high_risk = len([
        o for o in filtered
        if o["risk"] >= 70
    ])

    return render_template(
        "tracking.html",
        objects=filtered,
        total=len(filtered),
        satellites=satellites,
        debris=debris,
        high_risk=high_risk,
        search=search,
        obj_type=obj_type,
        risk_filter=risk_filter,
        sort_by=sort_by
    )

@app.route("/graph")
def graph():
    graph_data = build_graph()

    nodes = list(graph_data.nodes(data=True))
    edges = list(graph_data.edges(data=True))

    satellites = len([
        node for node, data in nodes
        if data["type"] == "satellite"
    ])

    debris = len([
        node for node, data in nodes
        if data["type"] == "debris"
    ])

    total_edges = len(edges)

    density = round(
        nx.density(graph_data),
        2
    )

    return render_template(
        "graph.html",
        nodes=nodes,
        edges=edges,
        satellites=satellites,
        debris=debris,
        total_edges=total_edges,
        density=density
    )

@app.route("/analytics")
def analytics():

    objects = get_objects()

    satellites = len([
        obj for obj in objects
        if obj["type"] == "satellite"
    ])

    debris = len([
        obj for obj in objects
        if obj["type"] == "debris"
    ])

    names = [
        obj["name"]
        for obj in objects
    ]

    risks = [
        obj["risk"]
        for obj in objects
    ]

    leo = len([
        obj for obj in objects
        if obj["altitude"] <= 2000
    ])

    meo = len([
        obj for obj in objects
        if 2000 < obj["altitude"] <= 35786
    ])

    geo = len([
        obj for obj in objects
        if obj["altitude"] > 35786
    ])

    average_risk = round(
        sum(risks) / len(risks),
        2
    )

    return render_template(
        "analytics.html",
        satellites=satellites,
        debris=debris,
        names=names,
        risks=risks,
        leo=leo,
        meo=meo,
        geo=geo,
        average_risk=average_risk
    )

    return render_template(
        "analytics.html",
        average_risk=average_risk
    )

@app.route("/api/analytics")
def api_analytics():
    objects = get_objects()

    satellites = len([o for o in objects if o["type"] == "satellite"])
    debris = len([o for o in objects if o["type"] == "debris"])

    names = [o["name"] for o in objects]
    risks = [o["risk"] for o in objects]

    leo = len([o for o in objects if o["altitude"] <= 2000])
    meo = len([o for o in objects if 2000 < o["altitude"] <= 35786])
    geo = len([o for o in objects if o["altitude"] > 35786])

    return jsonify({
        "names": names,
        "risks": risks,
        "satellites": satellites,
        "debris": debris,
        "leo": leo,
        "meo": meo,
        "geo": geo
    })

@app.route("/collision", methods=["GET", "POST"])
def collision():
    objects = get_objects()

    result = None
    probability_width = 0

    if request.method == "POST":
        object_name = request.form["object_name"]
        simulations = int(request.form["simulations"])

        result = run_monte_carlo(
            object_name,
            simulations
        )

        if result:
            probability_width = result["probability"]

    return render_template(
        "collision.html",
        objects=objects,
        result=result,
        probability_width=probability_width
    )

@app.route("/mission-planner", methods=["GET", "POST"])
def mission_planner():
    result = None

    if request.method == "POST":
        fuel = int(request.form["fuel"])

        result = optimize_mission(fuel)
        result["efficiency_width"] = min(
            100,
            result["efficiency"] * 20
        )

    return render_template(
        "mission_planner.html",
        result=result
    )

@app.route("/report")
def report():
    objects = get_objects()

    total = len(objects)

    satellites = len([
        o for o in objects
        if o["type"] == "satellite"
    ])

    debris = len([
        o for o in objects
        if o["type"] == "debris"
    ])

    average_risk = round(
        sum(o["risk"] for o in objects) / total,
        2
    )

    most_dangerous = max(
        objects,
        key=lambda o: o["risk"]
    )

    high_risk = [
        o for o in objects
        if o["risk"] >= 70
    ]

    leo = len([
        o for o in objects
        if o["altitude"] <= 2000
    ])

    meo = len([
        o for o in objects
        if 2000 < o["altitude"] <= 35786
    ])

    geo = len([
        o for o in objects
        if o["altitude"] > 35786
    ])

    recommendation = (
        f"Priorizar o monitoramento de {most_dangerous['name']}, "
        "pois apresenta o maior risco orbital."
        if most_dangerous["risk"] >= 70
        else "Cenário orbital dentro de níveis aceitáveis de segurança."
    )

    return render_template(
        "report.html",
        objects=objects,
        total=total,
        satellites=satellites,
        debris=debris,
        average_risk=average_risk,
        most_dangerous=most_dangerous,
        high_risk=high_risk,
        leo=leo,
        meo=meo,
        geo=geo,
        recommendation=recommendation
    )

@app.route("/report/export")
def export_report_csv():
    objects = get_objects()

    output = StringIO()
    writer = csv.writer(output)

    total = len(objects)
    satellites = len([o for o in objects if o["type"] == "satellite"])
    debris = len([o for o in objects if o["type"] == "debris"])
    average_risk = round(sum(o["risk"] for o in objects) / total, 2)
    high_risk = [o for o in objects if o["risk"] >= 70]
    most_dangerous = max(objects, key=lambda o: o["risk"])

    leo = len([o for o in objects if o["altitude"] <= 2000])
    meo = len([o for o in objects if 2000 < o["altitude"] <= 35786])
    geo = len([o for o in objects if o["altitude"] > 35786])

    writer.writerow(["ORBITALGUARD EXECUTIVE REPORT"])
    writer.writerow([])

    writer.writerow(["SYSTEM_INSIGHTS"])
    writer.writerow(["total_objects", total])
    writer.writerow(["total_satellites", satellites])
    writer.writerow(["total_debris", debris])
    writer.writerow(["average_risk", average_risk])
    writer.writerow(["high_risk_objects", len(high_risk)])
    writer.writerow(["leo_objects", leo])
    writer.writerow(["meo_objects", meo])
    writer.writerow(["geo_objects", geo])
    writer.writerow(["most_dangerous_object", most_dangerous["name"]])
    writer.writerow(["most_dangerous_risk", most_dangerous["risk"]])
    writer.writerow([])

    writer.writerow(["ALGORITHMS"])
    writer.writerow(["Dijkstra", "Caminho mínimo orbital"])
    writer.writerow(["Greedy", "Priorização de detritos"])
    writer.writerow(["Monte Carlo", "Simulação probabilística de colisão"])
    writer.writerow(["Dynamic Programming", "Otimização de combustível"])
    writer.writerow([])

    writer.writerow(["RECOMMENDATIONS"])
    if most_dangerous["risk"] >= 70:
        writer.writerow([
            f"Priorizar monitoramento/remoção de {most_dangerous['name']}"
        ])
    writer.writerow(["Executar simulações Monte Carlo periodicamente"])
    writer.writerow(["Reavaliar objetos em órbita LEO"])
    writer.writerow(["Usar Dijkstra para rotas orbitais seguras"])
    writer.writerow([])

    writer.writerow(["ORBITAL_OBJECTS"])
    writer.writerow([
        "id",
        "name",
        "type",
        "altitude_km",
        "risk",
        "risk_classification",
        "orbit_region",
        "recommendation"
    ])

    for obj in objects:
        risk = obj["risk"]
        altitude = obj["altitude"]

        if risk >= 70:
            risk_classification = "ALTO"
            recommendation = "Priorizar monitoramento/remocao"
        elif risk >= 40:
            risk_classification = "MEDIO"
            recommendation = "Monitorar com frequencia"
        else:
            risk_classification = "BAIXO"
            recommendation = "Monitoramento normal"

        if altitude <= 2000:
            orbit_region = "LEO"
        elif altitude <= 35786:
            orbit_region = "MEO"
        else:
            orbit_region = "GEO"

        writer.writerow([
            obj["id"],
            obj["name"],
            obj["type"],
            altitude,
            risk,
            risk_classification,
            orbit_region,
            recommendation
        ])

    response = Response(
        output.getvalue(),
        mimetype="text/csv"
    )

    response.headers["Content-Disposition"] = (
        "attachment; filename=orbitalguard_report.csv"
    )

    return response

@app.route("/report/export/excel")
def export_report_excel():
    objects = get_objects()

    total = len(objects)
    satellites = len([o for o in objects if o["type"] == "satellite"])
    debris = len([o for o in objects if o["type"] == "debris"])
    average_risk = round(sum(o["risk"] for o in objects) / total, 2)
    high_risk = [o for o in objects if o["risk"] >= 70]
    most_dangerous = max(objects, key=lambda o: o["risk"])

    leo = len([o for o in objects if o["altitude"] <= 2000])
    meo = len([o for o in objects if 2000 < o["altitude"] <= 35786])
    geo = len([o for o in objects if o["altitude"] > 35786])

    wb = Workbook()

    header_fill = PatternFill(
        start_color="00D4FF",
        end_color="00D4FF",
        fill_type="solid"
    )

    title_font = Font(
        bold=True,
        size=16
    )

    header_font = Font(
        bold=True,
        color="000000"
    )

    # DASHBOARD
    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = "OrbitalGuard Executive Report"
    ws["A1"].font = title_font

    dashboard_data = [
        ["Indicador", "Valor"],
        ["Total de objetos", total],
        ["Satélites", satellites],
        ["Detritos", debris],
        ["Risco médio", average_risk],
        ["Objetos de alto risco", len(high_risk)],
        ["LEO", leo],
        ["MEO", meo],
        ["GEO", geo],
        ["Objeto mais perigoso", most_dangerous["name"]],
        ["Risco do objeto mais perigoso", most_dangerous["risk"]]
    ]

    for row in dashboard_data:
        ws.append(row)

    # OBJECTS
    ws2 = wb.create_sheet("Objects")

    headers = [
        "ID",
        "Nome",
        "Tipo",
        "Altitude km",
        "Risco",
        "Classificação",
        "Região orbital",
        "Recomendação"
    ]

    ws2.append(headers)

    for obj in objects:
        risk = obj["risk"]
        altitude = obj["altitude"]

        if risk >= 70:
            classification = "ALTO"
            recommendation = "Priorizar monitoramento/remocao"
        elif risk >= 40:
            classification = "MEDIO"
            recommendation = "Monitorar com frequencia"
        else:
            classification = "BAIXO"
            recommendation = "Monitoramento normal"

        if altitude <= 2000:
            orbit = "LEO"
        elif altitude <= 35786:
            orbit = "MEO"
        else:
            orbit = "GEO"

        ws2.append([
            obj["id"],
            obj["name"],
            obj["type"],
            altitude,
            risk,
            classification,
            orbit,
            recommendation
        ])

    # RISK ANALYSIS
    ws3 = wb.create_sheet("Risk Analysis")

    ws3.append(["Objeto", "Risco", "Classificação"])

    for obj in sorted(objects, key=lambda o: o["risk"], reverse=True):
        risk = obj["risk"]

        if risk >= 70:
            classification = "ALTO"
        elif risk >= 40:
            classification = "MEDIO"
        else:
            classification = "BAIXO"

        ws3.append([
            obj["name"],
            risk,
            classification
        ])

    # INSIGHTS
    ws4 = wb.create_sheet("Executive Insights")

    ws4.append(["Insight", "Descrição"])

    ws4.append([
        "Objeto mais perigoso",
        f"{most_dangerous['name']} com risco {most_dangerous['risk']}"
    ])

    ws4.append([
        "Recomendação principal",
        f"Priorizar monitoramento/remoção de {most_dangerous['name']}"
        if most_dangerous["risk"] >= 70
        else "Cenário orbital dentro de níveis aceitáveis"
    ])

    ws4.append([
        "Dijkstra",
        "Usado para calcular caminho mínimo orbital"
    ])

    ws4.append([
        "Guloso",
        "Usado para priorizar detritos de maior risco"
    ])

    ws4.append([
        "Monte Carlo",
        "Usado para simular probabilidade de colisão"
    ])

    ws4.append([
        "Programação Dinâmica",
        "Usada para otimizar combustível e manobras"
    ])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[column].width = max_length + 4

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="orbitalguard_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@app.route("/route", methods=["GET", "POST"])
def route():
    objects = get_objects()

    result = None
    start = None
    end = None

    if request.method == "POST":
        start = request.form["start"]
        end = request.form["end"]

        result = find_best_route(
            start,
            end
        )

    return render_template(
        "route.html",
        objects=objects,
        result=result,
        start=start,
        end=end
    )

if __name__ == "__main__":
    app.run(debug=True)